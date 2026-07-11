import io
import os
import re
import textwrap

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

try:
    import xlsxwriter  # noqa: F401
    HAS_XLSXWRITER = True
except Exception:
    HAS_XLSXWRITER = False


st.set_page_config(page_title="CRISPR Indel / %OOF Explorer", layout="wide")
st.title("CRISPR Indel / %OOF Explorer")
st.caption(
    "In vitro: group = text before the last '-', dose = token after the last '-' (blank → 0). "
    "In vivo: only descriptions starting like A1-... or B2-... are treated as in vivo; "
    "group = leading letter, mouse = following digits, gRNA = text between the 1st and last '-', dose = last token (blank → 0)."
)


def load_sample() -> pd.DataFrame:
    return pd.DataFrame({
        "Sample_ID": ["py3A1", "py3A2", "py3A3", "py3B1", "py3B2", "py3B3", "py4A1", "py4A2", "py4A3", "py4B1", "py4B2", "py4B3"],
        "Description": [
            "A1-XXX-0.2", "A2-XXX-0.2", "A3-XXX-",
            "B1-YYY-0.2", "B2-YYY-0.2", "B3-YYY-0.2",
            "A1-XXX-0.4", "A2-XXX-0.4", "A3-XXX-0.4",
            "B1-YYY-0.4", "B2-YYY-0.4", "B3-YYY-0.4",
        ],
        "Amplicon_No": ["amp1"] * 12,
        "Indel%": [45.2, 47.1, 46.5, 10.1, 12.3, 9.8, 33.2, 32.5, 35.0, 55.0, 54.2, 53.5],
        "%OOF": [5.1, 4.8, 5.3, 0.9, 1.2, 1.0, 3.2, 3.5, 3.1, 8.0, 7.6, 7.9],
        "Reads_in_input": [1.2e5, 1.1e5, 1.3e5, 9.0e4, 8.5e4, 8.8e4, 2.0e5, 2.1e5, 1.8e5, 6.0e4, 6.5e4, 6.2e4],
        "Reads_aligned_all_amplicons": [1.0e5, 1.0e5, 1.1e5, 7.0e4, 6.8e4, 6.9e4, 1.6e5, 1.7e5, 1.5e5, 5.5e4, 5.6e4, 5.4e4],
    })


@st.cache_data
def read_csv(file) -> pd.DataFrame:
    return pd.read_csv(file)


def preview_text(df: pd.DataFrame, n: int = 20) -> str:
    if df is None or df.empty:
        return "<empty>"
    return df.head(n).to_csv(index=False)


def infer_columns(df: pd.DataFrame):
    cols = list(df.columns)
    amp_exact = "Amplicon_No" if "Amplicon_No" in cols else None
    desc_col = next((c for c in cols if re.search(r"^desc(ription)?$|description", c, flags=re.I)), None)
    sid_col = next((c for c in cols if re.search(r"^sample[_ ]?id$|sampleid|sample[-_ ]?name", c, flags=re.I)), None)
    indel_col = next((c for c in cols if re.search(r"\bindel\b", c, flags=re.I)), None)
    oof_col = next((c for c in cols if re.search(r"\boof\b|out[-_ ]?of[-_ ]?frame", c, flags=re.I)), None)
    rin_col = next((c for c in cols if re.search(r"reads.*input|input.*reads", c, flags=re.I)), None)
    raa_col = next((c for c in cols if re.search(r"reads.*aligned.*amplicon|aligned.*amplicon", c, flags=re.I)), None)
    amp_col = amp_exact or next((c for c in cols if re.search(r"\bamplicon(_?no)?\b|\bamplicon\b", c, flags=re.I)), None)
    if indel_col is None:
        indel_col = next((c for c in cols if re.search(r"indel.*%|%.*indel|indel[_ ]?percent", c, flags=re.I)), None)
    if oof_col is None:
        oof_col = next((c for c in cols if re.search(r"oof.*%|%.*oof|oof[_ ]?percent", c, flags=re.I)), None)
    return {
        "desc": desc_col,
        "sid": sid_col,
        "indel": indel_col,
        "oof": oof_col,
        "rin": rin_col,
        "raa": raa_col,
        "amp": amp_col,
    }


def numeric_from_string(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)
    s = str(x).strip()
    if not s:
        return np.nan
    s = s.replace("−", "-").replace("–", "-").replace(",", "")
    val = pd.to_numeric(s, errors="coerce")
    if not pd.isna(val):
        return float(val)
    m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
    return float(m.group(0)) if m else np.nan


def clean_metric_column(df, col):
    if col and col in df.columns:
        s = df[col]
        if pd.api.types.is_numeric_dtype(s):
            df[col] = pd.to_numeric(s, errors="coerce")
        else:
            df[col] = s.apply(numeric_from_string)
    return df


def clip_nonnegative_inplace(df, cols):
    for c in cols:
        if c and c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
            df[c] = df[c].mask(df[c] < 0, 0.0)


def n1(x):
    try:
        v = float(x)
        if v < 0 and abs(v) < 1e-12:
            v = 0.0
        return round(v, 1)
    except Exception:
        return np.nan


def parse_plate_well(sample_id: str):
    if sample_id is None or (isinstance(sample_id, float) and np.isnan(sample_id)):
        return "Unknown", "NA"
    s = str(sample_id).strip()
    m = re.match(r"^(?P<plate>.*?)[ _-]?(?P<well>[A-Ha-h](?:0?[1-9]|1[0-2]))$", s)
    if not m:
        return "Unknown", s
    plate = m.group("plate")
    well = m.group("well").upper()
    well = well[0] + str(int(well[1:]))
    return plate, well


def sort_doses_for_ui(dose_list):
    vals = [numeric_from_string(d) for d in dose_list]
    if all(pd.notna(v) for v in vals):
        return [d for d, _ in sorted(zip(dose_list, vals), key=lambda t: t[1], reverse=True)]
    return sorted(dose_list, key=lambda x: (str(x).lower(), str(x)))


def _is_oof_col(name: str) -> bool:
    return "oof" in str(name).lower()


def _guess_col_widths(df, min_w=6, max_w=28):
    widths = []
    for c in df.columns:
        header = str(c)
        body_max = max((len(str(v)) for v in df[c].fillna("").tolist()), default=0)
        w = max(len(header), body_max) + 2
        widths.append(max(min_w, min(w, max_w)))
    return widths


def _soft_wrap_desc(s: str, width: int = 24) -> str:
    s = str(s)
    for delim in ["-", "_"]:
        if delim in s and len(s) > width:
            parts = s.split(delim)
            lines, cur = [], ""
            for p in parts:
                add = (delim if cur else "") + p
                if len(cur) + len(add) > width and cur:
                    lines.append(cur)
                    cur = p
                else:
                    cur += add
            if cur:
                lines.append(cur)
            return "\n".join(lines)
    return "\n".join(textwrap.wrap(s, width=width)) if len(s) > width else s


def format_header_for_excel(name: str, desc_width: int = 24) -> str:
    s = str(name)
    m = re.match(r"^(?P<dose>.+?)-(Indel|OOF)-(?P<rep>rep\d+)$", s, flags=re.I)
    if m:
        left_wrapped = _soft_wrap_desc(m.group("dose"), width=desc_width)
        metric = m.group(2)
        return f"{left_wrapped}\n{metric}\n{m.group('rep')}"

    parts = s.split("_")
    if len(parts) >= 3 and re.fullmatch(r"(rep\d+|mean|SD)", parts[-1] or "", flags=re.I):
        metric = parts[-2]
        tail = parts[-1]
        left = "_".join(parts[:-2])
        left_wrapped = _soft_wrap_desc(left, width=desc_width)
        return f"{left_wrapped}\n{metric}\n{tail}"

    return _soft_wrap_desc(s, width=desc_width).replace("_", "\n").replace("-", "\n")


def style_sheet_xlsxwriter(writer, sheet_name, df):
    wb = writer.book
    ws = writer.sheets[sheet_name]
    base_fmt = {"bold": True, "text_wrap": True, "align": "center", "valign": "vcenter", "border": 1}
    hfmt_indel = wb.add_format({**base_fmt, "bg_color": "#F8D7DA"})
    hfmt_oof = wb.add_format({**base_fmt, "bg_color": "#D9E8FF"})

    for col, name in enumerate(df.columns):
        ws.write(0, col, format_header_for_excel(name), hfmt_oof if _is_oof_col(name) else hfmt_indel)

    narrow_sheets = {"type1_groups_full", "type1_groups_slim", "type2_doses", "group_gRNA_by_dose_reps"}
    max_cap = 10 if sheet_name in narrow_sheets else 16
    widths = _guess_col_widths(df, min_w=6, max_w=max_cap)
    for col, w in enumerate(widths):
        ws.set_column(col, col, w)
    ws.freeze_panes(1, 1)
    ws.autofilter(0, 0, len(df), len(df.columns) - 1)
    ws.set_row(0, 48)


def style_sheet_openpyxl(writer, sheet_name, df):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = writer.sheets[sheet_name]
    ws.freeze_panes = "A2"
    last_col = get_column_letter(df.shape[1])
    ws.auto_filter.ref = f"A1:{last_col}{df.shape[0] + 1}"
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    fill_indel = PatternFill("solid", fgColor="F8D7DA")
    fill_oof = PatternFill("solid", fgColor="D9E8FF")

    for j, name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j)
        cell.value = format_header_for_excel(name)
        cell.alignment = align
        cell.font = bold
        cell.border = border
        cell.fill = fill_oof if _is_oof_col(name) else fill_indel

    ws.row_dimensions[1].height = 48
    narrow_sheets = {"type1_groups_full", "type1_groups_slim", "type2_doses", "group_gRNA_by_dose_reps"}
    max_cap = 10 if sheet_name in narrow_sheets else 16
    widths = _guess_col_widths(df, min_w=6, max_w=max_cap)
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def style_sheet(writer, engine_name, sheet_name, df):
    if engine_name == "xlsxwriter":
        style_sheet_xlsxwriter(writer, sheet_name, df)
    else:
        style_sheet_openpyxl(writer, sheet_name, df)


def parse_ivv(description: str):
    if description is None or (isinstance(description, float) and np.isnan(description)):
        return "NON-MATCH", np.nan, "", "0", np.nan, False
    s = str(description).strip()
    if not s:
        return "NON-MATCH", np.nan, "", "0", np.nan, False

    parts = [p.strip() for p in s.split("-")]
    if len(parts) < 3:
        return "NON-MATCH", np.nan, "", "0", np.nan, False

    first = parts[0]
    m0 = re.match(r"^([A-Za-z])(\d+)$", first)
    if not m0:
        return "NON-MATCH", np.nan, "", "0", np.nan, False

    grp = m0.group(1).upper()
    rep_num = float(m0.group(2))
    middle = "-".join(parts[1:-1]).strip()
    if middle == "":
        return "NON-MATCH", np.nan, "", "0", np.nan, False

    dose_str = parts[-1].strip() or "0"
    dose_val = numeric_from_string(dose_str)
    return grp, rep_num, middle, dose_str, dose_val, True


def parse_invitro(description: str):
    if description is None or (isinstance(description, float) and np.isnan(description)):
        return "Unknown", "0", 0.0
    s = str(description).strip()
    if "-" in s:
        prefix, dose = s.rsplit("-", 1)
    else:
        prefix, dose = s, "0"
    if dose == "":
        dose = "0"
    return prefix, dose, numeric_from_string(dose)


def shade_map_for_doses(rank_order, palette):
    if not rank_order:
        return {}
    max_idx = len(palette) - 1
    return {dose: palette[min(i, max_idx)] for i, dose in enumerate(rank_order)}


def style_filled_bar_chart(fig):
    fig.update_traces(width=0.9, marker_line_width=0, opacity=1.0)
    fig.update_layout(bargap=0.12, bargroupgap=0.0)
    return fig


def render_indel_heatmap(df_plot: pd.DataFrame, metric_col: str, safe_mode: bool):
    st.subheader("Indel% heatmap")
    if safe_mode:
        st.info("Safe mode is on: heatmap is disabled.")
        return
    if metric_col not in df_plot.columns:
        st.info("Metric column not found.")
        return

    row_labels = list("ABCDEFGH")
    col_labels = [str(i) for i in range(1, 13)]
    row_index = {r: i for i, r in enumerate(row_labels)}
    col_index = {c: i for i, c in enumerate(col_labels)}

    def _well_to_rc(w):
        m = re.match(r"^\s*([A-Ha-h])\s*[-:]?\s*(\d{1,2})\s*$", str(w))
        if not m:
            return None
        return row_index[m.group(1).upper()], col_index[str(int(m.group(2)))]

    df_heat = df_plot[["_Plate", "_Well", metric_col]].copy()
    df_heat["_rc"] = df_heat["_Well"].map(_well_to_rc)
    df_heat = df_heat[df_heat["_rc"].notna()].copy()
    if df_heat.empty:
        st.caption("No valid wells.")
        return

    plates_for_heat = sorted(df_heat["_Plate"].dropna().unique().tolist())
    col_slots = st.columns(2)
    for i, plate in enumerate(plates_for_heat):
        sub = df_heat[df_heat["_Plate"] == plate].copy().drop_duplicates("_rc", keep="first")
        Z = np.full((8, 12), np.nan, dtype=float)
        text_matrix = np.empty((8, 12), dtype=object)
        text_matrix[:] = ""
        for _, r in sub.iterrows():
            rr, cc = r["_rc"]
            val = r[metric_col]
            if pd.notna(val):
                v = round(float(val))
                Z[rr, cc] = v
                text_matrix[rr, cc] = str(int(v))

        fig_hm = go.Figure(data=go.Heatmap(
            z=Z,
            x=col_labels,
            y=row_labels,
            colorscale="Viridis",
            zmin=0,
            zmax=100,
            colorbar=dict(title="Indel%", thickness=10, outlinewidth=0),
            text=text_matrix,
            texttemplate="%{text}",
            textfont=dict(color="black"),
            hovertemplate="Row %{y}, Col %{x}<br>Val: %{z:.0f}<extra></extra>",
        ))
        fig_hm.update_yaxes(autorange="reversed", tickfont=dict(color="black"))
        fig_hm.update_xaxes(side="top", tickfont=dict(color="black"))
        fig_hm.update_layout(
            title=dict(text=str(plate), x=0.9, xanchor="center", y=0.95, font=dict(color="black", size=20)),
            margin=dict(l=10, r=10, t=40, b=10),
            plot_bgcolor="white",
            paper_bgcolor="white",
            height=360,
        )
        with col_slots[i % 2]:
            st.plotly_chart(fig_hm, width="stretch", config={"displaylogo": False})
        if (i % 2) == 1 and (i + 1) < len(plates_for_heat):
            col_slots = st.columns(2)


safe_mode = False

with st.sidebar:
    st.header("1) Load data")
    uploaded = st.file_uploader("Upload CSV", type=["csv"])
    use_sample = st.checkbox("Use example dataset", value=not uploaded)
    st.markdown("**Trim low-quality rows**")
    apply_qc = st.checkbox("Enable QC trimming", value=True)
    qc_reads_min = st.number_input("Min Reads_in_input", min_value=0, value=500, step=50)
    qc_align_min = st.number_input("Min alignment%", min_value=0, max_value=100, value=80, step=1)
    st.header("Study type")
    in_vivo_mode = st.checkbox(
        "Analyze as in vivo data (A1-<gRNA>-<dose>)",
        value=False,
        help="Only descriptions starting with a letter plus digits, such as A1-... or B2-..., are treated as in vivo.",
    )

if uploaded is not None:
    raw = read_csv(uploaded)
elif use_sample:
    raw = load_sample()
    st.info("Using the built-in example dataset.")
else:
    st.stop()

info = infer_columns(raw)
if uploaded is not None and getattr(uploaded, "name", None):
    _input_name = uploaded.name
elif use_sample:
    _input_name = "example_dataset.csv"
else:
    _input_name = "input.csv"
_file_stem = os.path.splitext(os.path.basename(_input_name))[0]
excel_filename = f"{_file_stem}_For_GraphPad_Input.xlsx"

with st.expander("Detected columns / change if needed", expanded=False):
    c1, c2, c3 = st.columns(3)
    with c1:
        desc_col = st.selectbox("Description", options=raw.columns, index=(raw.columns.get_loc(info["desc"]) if info["desc"] in raw.columns else 0))
        sid_col = st.selectbox("Sample_ID", options=raw.columns, index=(raw.columns.get_loc(info["sid"]) if info["sid"] in raw.columns else 0))
    with c2:
        indel_col = st.selectbox("Indel% column", options=raw.columns, index=(raw.columns.get_loc(info["indel"]) if info["indel"] in raw.columns else 0))
        oof_col = st.selectbox("%OOF column", options=[None] + list(raw.columns), index=(1 + raw.columns.get_loc(info["oof"]) if info["oof"] in raw.columns else 0))
    with c3:
        rin_col = st.selectbox("Reads_in_input", options=[None] + list(raw.columns), index=(1 + raw.columns.get_loc(info["rin"]) if info["rin"] in raw.columns else 0))
        raa_col = st.selectbox("Reads_aligned_all_amplicons", options=[None] + list(raw.columns), index=(1 + raw.columns.get_loc(info["raa"]) if info["raa"] in raw.columns else 0))
    amp_col = st.selectbox("Amplicon column (optional)", options=[None] + list(raw.columns), index=(1 + raw.columns.get_loc(info["amp"]) if info["amp"] in raw.columns else 0))


df = raw.copy()
df = clean_metric_column(df, indel_col)
if oof_col:
    df = clean_metric_column(df, oof_col)
if rin_col:
    df[rin_col] = df[rin_col].apply(numeric_from_string)
if raa_col:
    df[raa_col] = df[raa_col].apply(numeric_from_string)
plates, wells = zip(*df[sid_col].map(parse_plate_well))
df["_Plate"] = list(plates)
df["_Well"] = list(wells)

if in_vivo_mode:
    gl, rn, grna, dstr, dval, ok = zip(*df[desc_col].map(parse_ivv))
    df["_Group"] = list(gl)
    df["_RepNum"] = list(rn)
    df["_gRNA"] = list(grna)
    df["_Dose"] = list(dstr)
    df["_DoseVal"] = list(dval)
    df["_InVivoMatch"] = list(ok)
    excluded_non_vivo_df = df.loc[~df["_InVivoMatch"]].copy()
    df = df[df["_InVivoMatch"]].copy()
    if df.empty:
        st.warning("In vivo mode is enabled, but none of the rows start with a treatment-group letter plus mouse number such as A1-<gRNA>-<dose>.")
        st.stop()
    if len(excluded_non_vivo_df) > 0:
        st.warning(
            f"In vivo mode kept {len(df)} matching row(s) and excluded {len(excluded_non_vivo_df)} non-matching row(s)."
        )
        with st.expander(f"Show rows excluded by in vivo parsing ({len(excluded_non_vivo_df)})", expanded=False):
            st.code(preview_text(excluded_non_vivo_df), language="text")
else:
    grp, dstr, dval = zip(*df[desc_col].map(parse_invitro))
    df["_Group"] = list(grp)
    df["_Dose"] = list(dstr)
    df["_DoseVal"] = list(dval)
    df["_RepNum"] = np.nan
    df["_gRNA"] = ""
    df["_InVivoMatch"] = True

df["_Dose"] = df["_Dose"].astype(str).replace({None: "0", "nan": "0", "NA": "0", "None": "0", "": "0"})
df["_DoseVal"] = df["_Dose"].apply(numeric_from_string)
if rin_col and raa_col:
    df["alignment%"] = np.where(df[rin_col] > 0, (df[raa_col] / df[rin_col]) * 100.0, np.nan)
else:
    df["alignment%"] = np.nan

if apply_qc:
    df_before_qc = df.copy()
    mask_qc = pd.Series(True, index=df.index)
    if rin_col:
        mask_qc &= df[rin_col].ge(qc_reads_min)
    if rin_col and raa_col:
        mask_qc &= df["alignment%"].ge(qc_align_min)
    removed_df = df_before_qc.loc[~mask_qc].copy()
    df = df.loc[mask_qc].copy()
    kept = len(df)
    total_pre_qc = len(df_before_qc)
    removed = total_pre_qc - kept
    st.success(f"QC trimming kept {kept}/{total_pre_qc} rows ({removed} removed).")
    with st.expander(f"Show filtered-out samples ({removed})", expanded=False):
        if removed > 0:
            st.code(preview_text(removed_df), language="text")
else:
    st.info("QC trimming is disabled.")

st.subheader("QC: Sequencing depth & alignment")
qc_cols_missing = []
if rin_col is None:
    qc_cols_missing.append("Reads_in_input")
if raa_col is None:
    qc_cols_missing.append("Reads_aligned_all_amplicons")
if qc_cols_missing:
    st.warning(f"QC plots require {', '.join(qc_cols_missing)}.")


def _on_plates_change():
    st.session_state.selected_amplicons = []
    st.session_state.selected_doses = []


def _on_amplicons_change():
    st.session_state.selected_doses = []


with st.sidebar:
    st.header("2) Filters")
    for k in ("selected_plates", "selected_amplicons", "selected_doses"):
        st.session_state.setdefault(k, [])

    available_plates = sorted(pd.Series(df["_Plate"].astype(str)).unique().tolist())
    st.write("**Plate(s)**")
    c1, c2 = st.columns(2)
    with c1:
        st.button("Select all", key="plates_all", on_click=lambda: st.session_state.update(selected_plates=available_plates, selected_amplicons=[], selected_doses=[]))
    with c2:
        st.button("Clear", key="plates_clear", on_click=lambda: st.session_state.update(selected_plates=[], selected_amplicons=[], selected_doses=[]))
    with st.expander("Choose plate(s)", expanded=False):
        q_pl = st.text_input("Search plates", key="plate_search", label_visibility="collapsed")
        plate_opts = [o for o in available_plates if not q_pl or q_pl.lower() in o.lower()]
        for o in plate_opts:
            st.checkbox(o, key=f"plate_opt_{o}", value=(o in st.session_state.selected_plates))
        if st.button("Apply plates", key="apply_plates"):
            st.session_state.selected_plates = [o for o in available_plates if st.session_state.get(f"plate_opt_{o}", False)]
            _on_plates_change()
    selected_plates = st.session_state["selected_plates"]
    st.caption("Selected: " + (", ".join(selected_plates) if selected_plates else "none"))

    if amp_col:
        amp_pool = df[df["_Plate"].astype(str).isin(selected_plates)][amp_col].astype(str) if selected_plates else df[amp_col].astype(str)
        available_amplicons = sorted([a for a in amp_pool.dropna().unique().tolist() if a != ""])
    else:
        available_amplicons = []
    st.write("**Amplicon**")
    amp_disabled = not selected_plates
    c1, c2 = st.columns(2)
    with c1:
        st.button("Select all", key="amp_all", disabled=amp_disabled, on_click=lambda: st.session_state.update(selected_amplicons=available_amplicons, selected_doses=[]))
    with c2:
        st.button("Clear", key="amp_clear", disabled=amp_disabled, on_click=lambda: st.session_state.update(selected_amplicons=[], selected_doses=[]))
    with st.expander("Choose amplicon(s)", expanded=False):
        q_amp = st.text_input("Search amplicons", key="amp_search", label_visibility="collapsed")
        amp_opts = [o for o in available_amplicons if not q_amp or q_amp.lower() in o.lower()]
        for o in amp_opts:
            st.checkbox(o, key=f"amp_opt_{o}", value=(o in st.session_state.selected_amplicons))
        if st.button("Apply amplicons", key="apply_amplicons"):
            st.session_state.selected_amplicons = [o for o in available_amplicons if st.session_state.get(f"amp_opt_{o}", False)]
            _on_amplicons_change()
    selected_amplicons = st.session_state["selected_amplicons"]
    st.caption("Selected: " + (", ".join(selected_amplicons) if selected_amplicons else "none"))

    if selected_plates:
        dose_pool = df[df["_Plate"].astype(str).isin(selected_plates)].copy()
        if amp_col and selected_amplicons:
            dose_pool = dose_pool[dose_pool[amp_col].astype(str).isin(selected_amplicons)]
        raw_doses = [d for d in dose_pool["_Dose"].astype(str).unique().tolist() if d != "NA"]
        available_doses = sort_doses_for_ui(raw_doses)
    else:
        available_doses = []
    st.write("**Dose(s)**")
    dose_disabled = not selected_plates
    c1, c2 = st.columns(2)
    with c1:
        st.button("Select all", key="dose_all", disabled=dose_disabled, on_click=lambda: st.session_state.update(selected_doses=available_doses))
    with c2:
        st.button("Clear", key="dose_clear", disabled=dose_disabled, on_click=lambda: st.session_state.update(selected_doses=[]))
    with st.expander("Choose dose(s)", expanded=False):
        q_dose = st.text_input("Search doses", key="dose_search", label_visibility="collapsed")
        dose_opts = [o for o in available_doses if not q_dose or q_dose.lower() in o.lower()]
        for o in dose_opts:
            st.checkbox(o, key=f"dose_opt_{o}", value=(o in st.session_state.selected_doses))
        if st.button("Apply doses", key="apply_doses"):
            st.session_state.selected_doses = [o for o in available_doses if st.session_state.get(f"dose_opt_{o}", False)]
    selected_doses = st.session_state["selected_doses"]
    st.caption("Selected: " + (", ".join(selected_doses) if selected_doses else "none"))

    st.header("3) Plot options")
    show_oof = st.checkbox("Show %OOF chart", value=False, key="show_oof")
    if in_vivo_mode:
        vivo_group_order = st.radio("Bar order", ["As input order", "By dose: High → Low", "By dose: Low → High"], index=0)
        group_by_series = True
        dose_order_choice = "High → Low"
        group_order_mode = "As input order"
        dose_for_group_sort = None
        ungrouped_order = "By mean ↓"
    else:
        group_by_series = st.checkbox("Group by dose series", value=True)
        dose_order_choice = st.radio("Dose order within group", options=["High → Low", "Low → High"], index=0)
        if group_by_series:
            group_order_mode = st.radio("Group order", options=["As input order", "By selected dose mean ↓", "By selected dose mean ↑"], index=0)
            dose_for_group_sort = st.selectbox("Dose for mean-based ordering", options=selected_doses, index=0) if group_order_mode != "As input order" and selected_doses else None
            ungrouped_order = "By mean ↓"
        else:
            group_order_mode = None
            dose_for_group_sort = None
            ungrouped_order = st.radio("Ungrouped bar order", options=["By mean ↓", "By mean ↑"], index=0)

    st.header("4) Image export")
    lock_ratio = st.checkbox("Lock aspect ratio 16:9", value=True)
    png_w = st.number_input("PNG width (px)", min_value=800, max_value=2400, value=1200, step=50)
    png_h = int(round(png_w * 9 / 16)) if lock_ratio else st.number_input("PNG height (px)", min_value=500, max_value=1500, value=675, step=25)
    png_scale = st.slider("Scale", 1, 3, 2)

    st.header("Troubleshooting")
    with st.expander("Troubleshooting", expanded=False):
        safe_mode = st.checkbox(
            "Safe mode (disable Plotly charts and Excel export)",
            value=False,
            help="Use this only if the app becomes unstable.",
        )

if not qc_cols_missing:
    if safe_mode:
        st.info("Safe mode is on: QC Plotly charts are disabled.")
    else:
        df["_ReadsIn_pos"] = np.where(df[rin_col] > 0, df[rin_col], np.nan)
        safe_max = np.nanmax(df["_ReadsIn_pos"]) if np.isfinite(np.nanmax(df["_ReadsIn_pos"])) else 1.0
        ticks = []
        t = 1.0
        while t <= max(1.0, safe_max * 1.2):
            ticks.append(t)
            t *= 10.0
        qc1, qc2 = st.columns(2)
        with qc1:
            fig1 = px.strip(df.assign(_X="All"), x="_X", y="_ReadsIn_pos", hover_data={sid_col: True, desc_col: True, rin_col: True})
            fig1.update_yaxes(type="log", tickvals=ticks, ticktext=[str(int(v)) if v >= 1 else str(v) for v in ticks])
            fig1.update_layout(xaxis_title="", yaxis_title="Reads_in_input (log scale)", showlegend=False, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig1, width="stretch")
        with qc2:
            fig2 = px.strip(df.assign(_X="All"), x="_X", y="alignment%", hover_data={sid_col: True, desc_col: True, rin_col: True, raa_col: True})
            fig2.update_layout(xaxis_title="", yaxis_title="alignment%", showlegend=False, margin=dict(l=10, r=10, t=10, b=10))
            st.plotly_chart(fig2, width="stretch")

st.markdown("---")

if not selected_plates:
    st.info("Select at least one plate to continue.")
    st.stop()
if not selected_doses:
    st.info("Select one or more doses to continue.")
    st.stop()

mask = df["_Plate"].astype(str).isin(selected_plates) & df["_Dose"].astype(str).isin(selected_doses)
if amp_col and selected_amplicons:
    mask &= df[amp_col].astype(str).isin(selected_amplicons)
fdf = df[mask].copy()
fdf_plot = fdf.copy()
clip_nonnegative_inplace(fdf_plot, [indel_col, oof_col, rin_col, raa_col, "alignment%"])

if in_vivo_mode:
    agg_indel = fdf_plot.groupby(["_Group", "_gRNA", "_Dose", "_DoseVal"], dropna=False)[indel_col].agg(mean="mean", std="std", count="count").reset_index()
    agg_indel["_DescLabel"] = agg_indel["_Group"].astype(str) + "-" + agg_indel["_gRNA"].astype(str) + "-" + agg_indel["_Dose"].astype(str)
else:
    agg_indel = fdf_plot.groupby(["_Group", "_Dose", "_DoseVal"], dropna=False)[indel_col].agg(mean="mean", std="std", count="count").reset_index()
    agg_indel["_DescLabel"] = agg_indel["_Group"].astype(str) + "-" + agg_indel["_Dose"].astype(str)

dose_low_to_high = agg_indel[["_Dose", "_DoseVal"]].drop_duplicates().sort_values(by=["_DoseVal", "_Dose"], ascending=[True, True])["_Dose"].astype(str).tolist()
dose_high_to_low = list(reversed(dose_low_to_high))

if in_vivo_mode:
    labels_input_order = list(pd.unique(fdf_plot["_Group"].astype(str) + "-" + fdf_plot["_gRNA"].astype(str) + "-" + fdf_plot["_Dose"].astype(str)))
    if vivo_group_order == "By dose: High → Low":
        dmap = agg_indel.set_index("_DescLabel")["_DoseVal"].to_dict()
        labels_sorted = sorted(labels_input_order, key=lambda x: (dmap.get(x, -np.inf),), reverse=True)
    elif vivo_group_order == "By dose: Low → High":
        dmap = agg_indel.set_index("_DescLabel")["_DoseVal"].to_dict()
        labels_sorted = sorted(labels_input_order, key=lambda x: (dmap.get(x, np.inf),))
    else:
        labels_sorted = labels_input_order
    x_categories = [lbl for lbl in labels_sorted if lbl in set(agg_indel["_DescLabel"].astype(str))]
else:
    groups_input_order = list(pd.unique(fdf_plot["_Group"].astype(str)))
    if group_by_series:
        if group_order_mode == "As input order" or not dose_for_group_sort:
            groups_sorted = groups_input_order
        else:
            by_group_dose = agg_indel[agg_indel["_Dose"].astype(str) == dose_for_group_sort].groupby("_Group")["mean"].mean()
            groups_with = [g for g in groups_input_order if g in by_group_dose.index]
            groups_without = [g for g in groups_input_order if g not in by_group_dose.index]
            desc = group_order_mode.endswith("↓")
            groups_with_sorted = sorted(groups_with, key=lambda g: by_group_dose[g], reverse=desc)
            groups_sorted = groups_with_sorted + groups_without
        within_order = dose_low_to_high if dose_order_choice == "Low → High" else dose_high_to_low
        x_categories = []
        for g in groups_sorted:
            present = [d for d in within_order if d in set(agg_indel.loc[agg_indel["_Group"] == g, "_Dose"].astype(str))]
            x_categories.extend([f"{g}-{d}" for d in present])
    else:
        asc = ungrouped_order == "By mean ↑"
        order_by_mean = agg_indel.sort_values(by=["mean", "_Group", "_DoseVal"], ascending=[asc, True, False])["_DescLabel"].astype(str).tolist()
        x_categories = list(dict.fromkeys(order_by_mean))

desc_vals = agg_indel["_DescLabel"].astype(str).tolist()
present = set(desc_vals)
x_categories = [x for x in x_categories if x in present] or list(dict.fromkeys(desc_vals))
agg_indel["_DescLabel"] = pd.Categorical(agg_indel["_DescLabel"], categories=x_categories, ordered=True)

group_shade_sets = [
    ["#7f0000", "#b30000", "#d7301f", "#ef6548", "#fc8d59", "#fdbb84"],
    ["#8c2d04", "#cc4c02", "#ec7014", "#fe9929", "#fec44f", "#fee391"],
    ["#08306b", "#08519c", "#2171b5", "#4292c6", "#6baed6", "#9ecae1"],
    ["#00441b", "#006d2c", "#238b45", "#41ab5d", "#74c476", "#a1d99b"],
    ["#3f007d", "#54278f", "#6a51a3", "#807dba", "#9e9ac8", "#bcbddc"],
    ["#004d4d", "#006d6f", "#008b8b", "#1aa3a3", "#66c2c2", "#99d8d8"],
    ["#000000", "#252525", "#525252", "#737373", "#969696", "#bdbdbd"],
]

color_discrete_map = {}
if in_vivo_mode:
    for gi, g in enumerate(pd.unique(agg_indel["_Group"].astype(str))):
        palette = group_shade_sets[gi % len(group_shade_sets)]
        sub = agg_indel[agg_indel["_Group"] == g]
        subu = sub.drop_duplicates(subset="_Dose").copy()
        if subu["_DoseVal"].notna().any():
            rank_order = subu.sort_values(["_DoseVal", "_Dose"], ascending=[False, True])["_Dose"].astype(str).tolist()
        else:
            rank_order = list(dict.fromkeys(sub["_Dose"].astype(str)))
        dose2shade = shade_map_for_doses(rank_order, palette)
        darkest = palette[0]
        for _, row in sub.iterrows():
            lbl = f"{row['_Group']}-{row['_gRNA']}-{row['_Dose']}"
            color_discrete_map[lbl] = dose2shade.get(str(row["_Dose"]), darkest)
else:
    for gi, (g, sub) in enumerate(agg_indel.groupby("_Group")):
        palette = group_shade_sets[gi % len(group_shade_sets)]
        subu = sub.drop_duplicates(subset="_Dose").copy()
        if subu["_DoseVal"].notna().any():
            rank_order = subu.sort_values(["_DoseVal", "_Dose"], ascending=[False, True])["_Dose"].astype(str).tolist()
        else:
            rank_order = [d for d in dose_high_to_low if d in set(subu["_Dose"].astype(str))]
            if not rank_order:
                rank_order = list(dict.fromkeys(subu["_Dose"].astype(str)))
        dose2shade = shade_map_for_doses(rank_order, palette)
        darkest = palette[0]
        for d in subu["_Dose"].astype(str).tolist():
            color_discrete_map[f"{g}-{d}"] = dose2shade.get(str(d), darkest)

if agg_indel.empty:
    st.warning("No data to plot.")
elif safe_mode:
    st.info("Safe mode is on: primary Plotly bar chart is disabled.")
else:
    fig = px.bar(
        agg_indel.sort_values("_DescLabel"),
        x="_DescLabel",
        y="mean",
        error_y="std",
        color="_DescLabel",
        color_discrete_map=color_discrete_map,
        hover_data={"_Group": True, "_Dose": True, "_DoseVal": True, "mean": ":.2f", "std": ":.2f", "count": True},
    )
    style_filled_bar_chart(fig)
    fig.update_layout(
        margin=dict(l=10, r=10, t=30, b=10),
        xaxis_title=None,
        yaxis_title="Mean Indel% (± SD)",
        showlegend=False,
        yaxis=dict(range=[0, 100], tickfont=dict(color="black"), title=dict(font=dict(color="black"))),
        xaxis=dict(tickfont=dict(color="black"), title=dict(font=dict(color="black"))),
        plot_bgcolor="white",
        paper_bgcolor="white",
    )
    config = {"toImageButtonOptions": {"format": "png", "filename": "indel_bars", "width": int(png_w), "height": int(png_h), "scale": int(png_scale)}, "displaylogo": False}
    st.plotly_chart(fig, width="stretch", config=config)

if show_oof and oof_col and oof_col in fdf_plot.columns:
    if in_vivo_mode:
        agg_oof = fdf_plot.groupby(["_Group", "_gRNA", "_Dose", "_DoseVal"], dropna=False)[oof_col].agg(mean="mean", std="std", count="count").reset_index()
        agg_oof["_DescLabel"] = agg_oof["_Group"].astype(str) + "-" + agg_oof["_gRNA"].astype(str) + "-" + agg_oof["_Dose"].astype(str)
    else:
        agg_oof = fdf_plot.groupby(["_Group", "_Dose", "_DoseVal"], dropna=False)[oof_col].agg(mean="mean", std="std", count="count").reset_index()
        agg_oof["_DescLabel"] = agg_oof["_Group"].astype(str) + "-" + agg_oof["_Dose"].astype(str)
    agg_oof["_DescLabel"] = pd.Categorical(agg_oof["_DescLabel"], categories=x_categories, ordered=True)
    if not agg_oof.empty:
        if safe_mode:
            st.info("Safe mode is on: %OOF Plotly bar chart is disabled.")
        else:
            fig2 = px.bar(
                agg_oof.sort_values("_DescLabel"),
                x="_DescLabel",
                y="mean",
                error_y="std",
                color="_DescLabel",
                color_discrete_map=color_discrete_map,
                hover_data={"_Group": True, "_Dose": True, "_DoseVal": True, "mean": ":.2f", "std": ":.2f", "count": True},
            )
            style_filled_bar_chart(fig2)
            fig2.update_layout(
                margin=dict(l=10, r=10, t=30, b=10),
                xaxis_title=None,
                yaxis_title="Mean %OOF (± SD)",
                showlegend=False,
                yaxis=dict(range=[0, 100], tickfont=dict(color="black"), title=dict(font=dict(color="black"))),
                xaxis=dict(tickfont=dict(color="black"), title=dict(font=dict(color="black"))),
                plot_bgcolor="white",
                paper_bgcolor="white",
            )
            config2 = {"toImageButtonOptions": {"format": "png", "filename": "oof_bars", "width": int(png_w), "height": int(png_h), "scale": int(png_scale)}, "displaylogo": False}
            st.plotly_chart(fig2, width="stretch", config=config2)

ordered_cols = ["Sample_ID", "Description", "Indel%", "%OOF", "Reads_in_input", "Reads_aligned_all_amplicons", "alignment%"]
col_map = {
    "Sample_ID": sid_col,
    "Description": desc_col,
    "Indel%": indel_col,
    "%OOF": oof_col,
    "Reads_in_input": rin_col,
    "Reads_aligned_all_amplicons": raa_col,
    "alignment%": "alignment%",
}
rows_df = pd.DataFrame({k: (np.nan if col_map[k] is None else fdf[col_map[k]]) for k in ordered_cols})
clip_nonnegative_inplace(rows_df, ["Indel%", "%OOF", "Reads_in_input", "Reads_aligned_all_amplicons", "alignment%"])

if in_vivo_mode:
    def _label_to_group_grna(lbl: str):
        parts = str(lbl).split("-")
        if len(parts) < 3:
            return "", ""
        return parts[0].strip(), "-".join(parts[1:-1]).strip()

    pair_order_plot = []
    for lbl in x_categories:
        key = _label_to_group_grna(lbl)
        if key not in pair_order_plot:
            pair_order_plot.append(key)

    seen = set()
    pair_order_input = []
    for _, row in fdf_plot.iterrows():
        key = (str(row["_Group"]), str(row["_gRNA"]))
        if key not in seen:
            seen.add(key)
            pair_order_input.append(key)

    dose_columns_order = dose_low_to_high if vivo_group_order == "By dose: Low → High" else dose_high_to_low

    def _rep_sort_key(sub):
        if "_RepNum" in sub.columns and sub["_RepNum"].notna().any():
            return sub.sort_values(by=["_RepNum", sid_col])
        if "_Well" in sub.columns and sub["_Well"].notna().any():
            return sub.sort_values(by=["_Well", sid_col])
        return sub.sort_values(by=[sid_col])

    max_reps = int(fdf_plot.groupby(["_Group", "_gRNA", "_Dose"]).size().max()) if not fdf_plot.empty else 1
    cols = []
    for d in dose_columns_order:
        for i in range(1, max_reps + 1):
            cols.append(f"{d}-Indel-rep{i}")
    if oof_col and oof_col in fdf_plot.columns:
        for d in dose_columns_order:
            for i in range(1, max_reps + 1):
                cols.append(f"{d}-OOF-rep{i}")

    def _is_pbs(d):
        ds = str(d).strip().lower()
        return ds in {"", "0", "0.0", "pbs"}

    nonpbs_rows, pbs_rows = [], []
    for G, g in pair_order_plot:
        doses_here = pd.unique(fdf_plot[(fdf_plot["_Group"] == G) & (fdf_plot["_gRNA"] == g)]["_Dose"].astype(str))
        has_pbs = any(_is_pbs(d) for d in doses_here)
        if has_pbs:
            row_pbs = {"gRNA": f"PBS-{G}-{g}"}
            for d in dose_columns_order:
                sub = fdf_plot[(fdf_plot["_Group"] == G) & (fdf_plot["_gRNA"] == g) & (fdf_plot["_Dose"].astype(str) == str(d))]
                if _is_pbs(d):
                    sub = _rep_sort_key(sub)
                    vals_indel = [n1(v) for v in sub[indel_col].tolist()]
                    for i in range(1, max_reps + 1):
                        row_pbs[f"{d}-Indel-rep{i}"] = vals_indel[i - 1] if i - 1 < len(vals_indel) else np.nan
                    if oof_col and oof_col in fdf_plot.columns:
                        vals_oof = [n1(v) for v in sub[oof_col].tolist()]
                        for i in range(1, max_reps + 1):
                            row_pbs[f"{d}-OOF-rep{i}"] = vals_oof[i - 1] if i - 1 < len(vals_oof) else np.nan
                else:
                    for i in range(1, max_reps + 1):
                        row_pbs[f"{d}-Indel-rep{i}"] = np.nan
                        if oof_col and oof_col in fdf_plot.columns:
                            row_pbs[f"{d}-OOF-rep{i}"] = np.nan
            pbs_rows.append(((G, g), row_pbs))

        has_nonpbs = any((not _is_pbs(d)) and (not fdf_plot[(fdf_plot["_Group"] == G) & (fdf_plot["_gRNA"] == g) & (fdf_plot["_Dose"].astype(str) == str(d))].empty) for d in dose_columns_order)
        if has_nonpbs:
            row = {"gRNA": f"{G}-{g}"}
            for d in dose_columns_order:
                sub = fdf_plot[(fdf_plot["_Group"] == G) & (fdf_plot["_gRNA"] == g) & (fdf_plot["_Dose"].astype(str) == str(d))]
                if _is_pbs(d):
                    for i in range(1, max_reps + 1):
                        row[f"{d}-Indel-rep{i}"] = np.nan
                        if oof_col and oof_col in fdf_plot.columns:
                            row[f"{d}-OOF-rep{i}"] = np.nan
                    continue
                sub = _rep_sort_key(sub)
                vals_indel = [n1(v) for v in sub[indel_col].tolist()]
                for i in range(1, max_reps + 1):
                    row[f"{d}-Indel-rep{i}"] = vals_indel[i - 1] if i - 1 < len(vals_indel) else np.nan
                if oof_col and oof_col in fdf_plot.columns:
                    vals_oof = [n1(v) for v in sub[oof_col].tolist()]
                    for i in range(1, max_reps + 1):
                        row[f"{d}-OOF-rep{i}"] = vals_oof[i - 1] if i - 1 < len(vals_oof) else np.nan
            nonpbs_rows.append(row)

    pbs_rows_sorted = []
    for key in pair_order_input:
        for k2, row in pbs_rows:
            if key == k2:
                pbs_rows_sorted.append(row)

    gp_table = pd.DataFrame(nonpbs_rows + pbs_rows_sorted, columns=["gRNA"] + cols)
    st.subheader("Download")
    if safe_mode:
        st.info("Safe mode is on: Excel export is disabled.")
    else:
        excel_engine = "xlsxwriter" if HAS_XLSXWRITER else "openpyxl"
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine=excel_engine) as writer:
            gp_table.to_excel(writer, index=False, sheet_name="group_gRNA_by_dose_reps")
            rows_df.to_excel(writer, index=False, sheet_name="rows_used")
            style_sheet(writer, excel_engine, "group_gRNA_by_dose_reps", gp_table)
            style_sheet(writer, excel_engine, "rows_used", rows_df)
        st.download_button("Download Excel (.xlsx)", data=buf.getvalue(), file_name=excel_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    render_indel_heatmap(fdf_plot, indel_col, safe_mode)

    st.subheader("gRNA table (In vivo)")
    st.code(preview_text(gp_table), language="text")
else:
    groups_in_order = []
    for lbl in x_categories:
        g = str(lbl).rsplit("-", 1)[0]
        if g not in groups_in_order:
            groups_in_order.append(g)
    doses_high_to_low = list(reversed(dose_low_to_high))
    gp_key = "_Group"
    src_df = fdf_plot.copy()
    if "_Well" in src_df.columns and src_df["_Well"].notna().any():
        fdf_sorted = src_df.sort_values(by=[gp_key, "_Dose", "_Well"])
    else:
        fdf_sorted = src_df.sort_values(by=[gp_key, "_Dose", sid_col])
    fdf_sorted["_rep_idx"] = fdf_sorted.groupby([gp_key, "_Dose"]).cumcount() + 1

    def make_type1(metric_label, metric_col):
        max_reps_global = 1
        for d in doses_high_to_low:
            cnt = fdf_sorted[fdf_sorted["_Dose"].astype(str) == d].groupby(gp_key)[metric_col].count()
            if not cnt.empty:
                max_reps_global = max(max_reps_global, int(cnt.max()))
        cols = []
        for d in doses_high_to_low:
            cols += [f"{d}_{metric_label}_rep{i}" for i in range(1, max_reps_global + 1)]
            cols += [f"{d}_{metric_label}_mean", f"{d}_{metric_label}_SD"]
        rows = []
        for g in groups_in_order:
            row = {"Description": g}
            subg = fdf_sorted[fdf_sorted[gp_key] == g]
            for d in doses_high_to_low:
                subgd = subg[subg["_Dose"].astype(str) == d]
                vals = [n1(v) for v in subgd.sort_values(["_rep_idx"])[metric_col].tolist()]
                for i in range(1, max_reps_global + 1):
                    row[f"{d}_{metric_label}_rep{i}"] = vals[i - 1] if i - 1 < len(vals) else np.nan
                if subgd.empty or metric_col not in subgd.columns:
                    row[f"{d}_{metric_label}_mean"] = np.nan
                    row[f"{d}_{metric_label}_SD"] = np.nan
                else:
                    vals_raw = subgd[metric_col].astype(float)
                    row[f"{d}_{metric_label}_mean"] = n1(vals_raw.mean())
                    row[f"{d}_{metric_label}_SD"] = n1(vals_raw.std(ddof=1))
            rows.append(row)
        out = pd.DataFrame(rows, columns=["Description"] + cols)
        out["Description"] = pd.Series(out.get("Description", pd.Series(dtype="object")), dtype="object")
        return out

    def make_type2(metric_label, metric_col):
        max_reps_global = 1
        for g in groups_in_order:
            cnt = fdf_sorted[fdf_sorted[gp_key] == g].groupby("_Dose")[metric_col].count()
            if not cnt.empty:
                max_reps_global = max(max_reps_global, int(cnt.max()))
        cols = [f"{g}_{metric_label}_rep{i}" for g in groups_in_order for i in range(1, max_reps_global + 1)]
        rows = []
        for d in doses_high_to_low:
            row = {"Dose": d}
            subd = fdf_sorted[fdf_sorted["_Dose"].astype(str) == d]
            for g in groups_in_order:
                subdg = subd[subd[gp_key] == g]
                vals = [n1(v) for v in subdg.sort_values(["_rep_idx"])[metric_col].tolist()]
                for i in range(1, max_reps_global + 1):
                    row[f"{g}_{metric_label}_rep{i}"] = vals[i - 1] if i - 1 < len(vals) else np.nan
            rows.append(row)
        out = pd.DataFrame(rows, columns=["Dose"] + cols)
        out["Dose"] = pd.Series(out.get("Dose", pd.Series(dtype="object")), dtype="object")
        return out

    metrics_for_export = [{"name": "Indel%", "col": indel_col}]
    if oof_col and oof_col in fdf_plot.columns:
        metrics_for_export.append({"name": "%OOF", "col": oof_col})

    type1_parts = [pd.DataFrame({"Description": pd.Series(groups_in_order, dtype="object")}).set_index("Description")]
    type2_parts = [pd.DataFrame({"Dose": pd.Series(doses_high_to_low, dtype="object")}).set_index("Dose")]
    for m in metrics_for_export:
        t1_piece = make_type1(m["name"], m["col"]).copy()
        t2_piece = make_type2(m["name"], m["col"]).copy()
        if "Description" in t1_piece.columns:
            type1_parts.append(t1_piece.set_index("Description"))
        if "Dose" in t2_piece.columns:
            type2_parts.append(t2_piece.set_index("Dose"))
    type1_full_df = pd.concat(type1_parts, axis=1).reset_index()
    type2_df = pd.concat(type2_parts, axis=1).reset_index()
    type1_slim_df = type1_full_df[[c for c in type1_full_df.columns if not (str(c).endswith("_mean") or str(c).endswith("_SD"))]].copy()

    st.subheader("Download")
    if safe_mode:
        st.info("Safe mode is on: Excel export is disabled.")
    else:
        excel_engine = "xlsxwriter" if HAS_XLSXWRITER else "openpyxl"
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine=excel_engine) as writer:
            if not type1_slim_df.empty:
                type1_slim_df.to_excel(writer, index=False, sheet_name="type1_groups_slim")
                style_sheet(writer, excel_engine, "type1_groups_slim", type1_slim_df)
            if not type1_full_df.empty:
                type1_full_df.to_excel(writer, index=False, sheet_name="type1_groups_full")
                style_sheet(writer, excel_engine, "type1_groups_full", type1_full_df)
            if not type2_df.empty:
                type2_df.to_excel(writer, index=False, sheet_name="type2_doses")
                style_sheet(writer, excel_engine, "type2_doses", type2_df)
            rows_df.to_excel(writer, index=False, sheet_name="rows_used")
            style_sheet(writer, excel_engine, "rows_used", rows_df)
        st.download_button("Download Excel (.xlsx)", data=buf.getvalue(), file_name=excel_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    render_indel_heatmap(fdf_plot, indel_col, safe_mode)

    st.subheader("In vitro tables (Preview)")
    which_gp = st.radio("Layout", options=["Type 1 slim (Rows=Group)", "Type 1 full (Rows=Group)", "Type 2 (Rows=Dose)"], index=0, horizontal=True)
    preview_df = type1_slim_df if which_gp.startswith("Type 1 slim") else type1_full_df if which_gp.startswith("Type 1 full") else type2_df
    st.code(preview_text(preview_df), language="text")

st.subheader("Rows used")
st.code(preview_text(rows_df), language="text")
